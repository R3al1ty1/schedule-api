from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, Header
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi.responses import StreamingResponse
import openpyxl
from io import BytesIO

from core.db_helper import db_helper
from core.models import booking as booking_model
from core.utils import verify_admin
from telegram_bot.utils.utils import send_excel_file
from fastapi.responses import JSONResponse


router = APIRouter(tags=["Schedule"])

db = db_helper.session_getter


@router.get("/export/excel/", response_class=StreamingResponse)
async def export_schedule_to_excel(
    user_id: int = Header(...),
    session: AsyncSession = Depends(db)
):
    """Экспорт расписания в Excel файл."""
    # Получаем текущую дату
    current_date = datetime.now()
    
    # Рассчитываем даты для фильтрации (6 месяцев назад и 6 месяцев вперед)
    start_date = current_date - timedelta(days=180)
    end_date = current_date + timedelta(days=180)
    
    # Получаем все бронирования за указанный период
    query = select(booking_model.Booking).where(
        booking_model.Booking.start_date >= start_date,
        booking_model.Booking.end_date <= end_date
    ).order_by(booking_model.Booking.start_date, booking_model.Booking.end_date)
    
    result = await session.execute(query)
    bookings = result.scalars().all()
    
    # Создаем новый Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расписание"
    
    from openpyxl.styles import Border, Side, PatternFill, Font
    
    # Определяем стили
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bottom_border = Border(bottom=Side(style='medium'))
    right_border = Border(right=Side(style='medium'))
    corner_border = Border(
        right=Side(style='medium'),
        bottom=Side(style='medium')
    )

    headers = [
        "Номер", "ID пользователя", "Дата начала", "Дата окончания", "Количество человек",
        "Название", "Тема мероприятия", "Описание", "Статус заявки", "Целевая аудитория" ,
        "Тип регистрации", "Логистика участников", "Тип программы", "Место", "Размещение участников",
        "Количество экспертов", "ФИО куратора", "Должность куратора",
        "Контакты куратора", "Дополнительная информация"
    ]
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = bottom_border
        if col == 1:
            cell.border = corner_border
    
    for row, booking in enumerate(bookings, 2):
        cell = ws.cell(row=row, column=1, value=booking.id)
        cell.border = right_border
        
        ws.cell(row=row, column=2, value=booking.user_id)
        ws.cell(row=row, column=3, value=booking.start_date.strftime("%d.%m.%Y"))
        ws.cell(row=row, column=4, value=booking.end_date.strftime("%d.%m.%Y"))
        ws.cell(row=row, column=5, value=booking.people_count)
        ws.cell(row=row, column=6, value=booking.name)
        ws.cell(row=row, column=7, value=booking.theme)
        ws.cell(row=row, column=8, value=booking.description)
        ws.cell(row=row, column=9, value=booking.status)
        ws.cell(row=row, column=10, value=booking.target_audience)
        ws.cell(row=row, column=11, value=booking.registration)
        ws.cell(row=row, column=12, value=booking.logistics)
        ws.cell(row=row, column=13, value=booking.type)
        ws.cell(row=row, column=14, value=booking.place)
        ws.cell(row=row, column=15, value=booking.participants_accomodation)
        ws.cell(row=row, column=16, value=booking.experts_count)
        ws.cell(row=row, column=17, value=booking.curator_fio)
        ws.cell(row=row, column=18, value=booking.curator_position)
        ws.cell(row=row, column=19, value=booking.curator_contact)
        ws.cell(row=row, column=20, value=booking.other_info)
    
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2
    
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    await send_excel_file(
        user_id=user_id,
        file=excel_file,
        filename=f"Расписание_{current_date.strftime('%d_%m_%Y')}.xlsx"
    )
    
    return JSONResponse(
        content={"status": "success", "message": "Файл успешно отправлен в Telegram"},
        status_code=200
    )

    # return StreamingResponse(
    #     excel_file,
    #     media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    #     headers={
    #         "Content-Disposition": f"attachment; filename=schedule_{current_date.strftime('%Y%m%d')}.xlsx"
    #     }
    # )
